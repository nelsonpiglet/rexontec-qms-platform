"""
REXONTEC OQC — 成檢表 Excel 解析器
支援 MD1003RX 格式（力山公司成品檢驗表）

Excel 欄位對應（1-indexed）：
  Col A (1)  : 類別字元（垂直書寫 section 名稱）
  Col B (2)  : 項次（1, 2, 3 ...）
  Col C (3)  : 檢驗項目（merged C:J）
  Col K (11) : 規格標準
  Col L (12) : 工具/手法
  Col M (13) : 等級（CR/MA/MI）
  Col N+ (14+): 樣品量測值（OK/NG 或數值）

Section 識別：Col A 垂直字元跨越 item row 及之後幾列
Item 識別：Col B 為數字且 Col C 有內容
Type 識別：若樣品值全為 OK/NG → pf；若樣品值為純數字 → num；其餘 → pf
"""
import io
import re


def _cell_val(ws, row: int, col: int, merged_map: dict) -> str:
    """取儲存格值，自動從 merged_map 補 None 值。"""
    v = ws.cell(row, col).value
    if v is None:
        v = merged_map.get((row, col))
    if v is None:
        return ""
    return str(v).strip()


def _build_merged_map(ws) -> dict:
    """建立合併儲存格查找表 (row, col) → value（以左上角值填入）。"""
    m = {}
    for rng in ws.merged_cells.ranges:
        top_val = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    m[(r, c)] = top_val
    return m


def _is_numeric(s: str) -> bool:
    """判斷字串是否為純數值（不含 '/'）。"""
    try:
        float(s.replace(",", "").replace("，", ""))
        return True
    except (ValueError, TypeError):
        return False


def _detect_type(samples: list[str]) -> str:
    """
    依樣品值判斷檢驗類型：
    - 若非空樣品皆為 OK/NG → 'pf'
    - 若非空樣品多數為純數字 → 'num'
    - 其餘 → 'pf'
    """
    non_empty = [s for s in samples if s]
    if not non_empty:
        return "pf"
    pf_set = {"ok", "ng", "pass", "fail", "o", "x", "合格", "不合格"}
    pf_count = sum(1 for s in non_empty if s.lower().strip() in pf_set)
    num_count = sum(1 for s in non_empty if _is_numeric(s))
    if pf_count == len(non_empty):
        return "pf"
    if num_count >= len(non_empty) * 0.7:
        return "num"
    return "pf"


def _normalize_grade(raw: str) -> str:
    """將各種等級寫法統一為 CR / MA / MI。"""
    r = raw.strip().upper()
    if "CR" in r or "嚴重" in r or "致命" in r:
        return "CR"
    if "MI" in r or "輕微" in r or "次要" in r:
        return "MI"
    return "MA"  # default


def _parse_spec_range(spec: str) -> tuple:
    """
    從規格字串解析 (unit, min, max)。
    支援格式：
      '400mm ↑'     → unit='mm', min=400, max=None
      '20MΩ ↑'     → pf（不解析，回傳 '',None,None）
      '0.07ohm±5%\\n(0.0665~0.0735)' → unit='ohm', min=0.0665, max=0.0735
      '69.6761±2.5%mV.s/rad'  → unit='mV.s/rad', min=67.9, max=71.4
      '3700～3900RPM'  → unit='RPM', min=3700, max=3900
    """
    if not spec:
        return "", None, None

    # 若規格含換行，優先取括號內 (min~max)
    if "\n" in spec:
        # 找 (X~Y) 或 (X～Y)
        m = re.search(r"\(([0-9.]+)\s*[~～]\s*([0-9.]+)\)", spec)
        if m:
            try:
                mn, mx = float(m.group(1)), float(m.group(2))
                # unit from first line
                first_line = spec.split("\n")[0]
                unit = re.sub(r"[0-9.±%\s↑↓≦≧≤≥<>()]+", "", first_line).strip()
                unit = re.sub(r"^[oO]hm", "ohm", unit)
                return unit, mn, mx
            except ValueError:
                pass

    # ±P% 格式
    m = re.search(r"([0-9.]+)\s*([A-Za-zΩμ./·%一-鿿]*)\s*±\s*([0-9.]+)\s*%", spec)
    if m:
        try:
            center = float(m.group(1))
            unit = m.group(2).strip() or ""
            pct = float(m.group(3)) / 100
            # strip leftover unit chars from trailing
            trailing = re.sub(r"^[0-9.±%\s↑↓≦≧≤≥<>()\n]+", "", spec[m.end():]).strip()
            if trailing:
                unit = (unit + trailing).strip()
            return unit, round(center * (1 - pct), 6), round(center * (1 + pct), 6)
        except ValueError:
            pass

    # X~Y unit 格式
    m = re.search(r"([0-9.]+)\s*[~～]\s*([0-9.]+)\s*([A-Za-zΩμ./·%一-鿿]*)", spec)
    if m:
        try:
            unit = m.group(3).strip()
            return unit, float(m.group(1)), float(m.group(2))
        except ValueError:
            pass

    # X unit ↑ / ↓ 格式
    m = re.search(r"([0-9.]+)\s*([A-Za-zΩμ./·%一-鿿]+)?\s*([↑↓]|↑|↓|≦|≧|≤|≥)", spec)
    if m:
        try:
            val = float(m.group(1))
            unit = (m.group(2) or "").strip()
            direction = m.group(3)
            if direction in ("↑", "≧", "≥"):
                return unit, val, None
            else:
                return unit, None, val
        except ValueError:
            pass

    # 無法解析，回傳空
    return "", None, None


def parse_oqc_excel(file_obj) -> list:
    """
    解析 OQC 成檢表 Excel（MD1003RX 格式），
    回傳 sections list（與 inspection_data.py sections 格式相容）。

    Parameters
    ----------
    file_obj : bytes / BytesIO / 檔案路徑 str

    Returns
    -------
    list of section dicts：
      [{"id": "A", "label": "包裝類 01", "subtitle": "", "items": [...]}, ...]
    """
    import openpyxl

    if isinstance(file_obj, (bytes, bytearray)):
        file_obj = io.BytesIO(file_obj)

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    merged_map = _build_merged_map(ws)

    def cv(r, c):
        return _cell_val(ws, r, c, merged_map)

    max_row = ws.max_row
    max_col = ws.max_column

    # ── 收集所有 item rows（col B 為數字且 col C 有文字）──────────────
    item_rows = []
    for row in range(1, max_row + 1):
        col_b = cv(row, 2)
        col_c = cv(row, 3)
        if col_b.isdigit() and 1 <= int(col_b) <= 30 and col_c:
            # 收集樣品值（col N=14 往後最多 30 欄）
            samples = [
                cv(row, c)
                for c in range(14, min(14 + 30, max_col + 1))
                if cv(row, c)
            ]
            item_rows.append({
                "row":      row,
                "no":       int(col_b),
                "name":     col_c,
                "spec":     cv(row, 11),
                "tool":     cv(row, 12),
                "grade":    _normalize_grade(cv(row, 13)),
                "samples":  samples,
            })

    if not item_rows:
        return []

    # ── 分組成 sections（item 編號重置 = 新 section）─────────────────
    groups: list[list[dict]] = []
    cur: list[dict] = []
    prev_no = 0

    for item in item_rows:
        # item.no 重置到 1 且 cur 不空 → 新 section
        if cur and item["no"] <= prev_no:
            groups.append(cur)
            cur = []
        cur.append(item)
        prev_no = item["no"]

    if cur:
        groups.append(cur)

    # ── 為每組重建 section 標題（從 col A 垂直字元）──────────────────
    def extract_section_label(items: list[dict]) -> str:
        """從 col A 垂直字元還原 section 標籤。
        最多看到 last_item_row + 2，避免讀入下一 section 的字元。"""
        first_row = items[0]["row"]
        last_row  = items[-1]["row"]
        chars = []
        for row in range(first_row, last_row + 3):  # +2 offset
            if row > max_row:
                break
            v = cv(row, 1)
            if v and v not in chars:
                chars.append(v)
        # 清理多餘空格，保留中文與數字
        label = "".join(chars)
        return label

    # ── 組建最終 sections ─────────────────────────────────────────
    SECTION_IDS = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
    result = []

    for idx, group in enumerate(groups):
        sec_id    = SECTION_IDS[idx] if idx < len(SECTION_IDS) else str(idx + 1)
        raw_label = extract_section_label(group)
        # 嘗試從標籤提取工具/副標
        subtitle  = ""
        if group[0]["tool"]:
            subtitle = group[0]["tool"]

        items_out = []
        for item in group:
            itype = _detect_type(item["samples"])

            # 清理 spec：取第一行（去除換行後說明文字）
            spec_clean = (item["spec"] or "目視合格").split("\n")[0].strip()

            out = {
                "id":    f"{sec_id}{item['no']}",
                "no":    f"{item['no']}.0",
                "name":  item["name"],
                "spec":  spec_clean,
                "grade": item["grade"],
                "type":  itype,
                "tool":  item["tool"] or "目視",
            }

            if itype == "num":
                unit, mn, mx = _parse_spec_range(item["spec"])
                out["unit"] = unit
                out["min"]  = mn
                out["max"]  = mx

            items_out.append(out)

        result.append({
            "id":       sec_id,
            "label":    raw_label if raw_label else f"Section {idx + 1}",
            "subtitle": subtitle,
            "items":    items_out,
        })

    return result


def extract_header_meta(file_obj) -> dict:
    """
    嘗試從 Excel 表頭解析機種、日期、客戶等資訊。
    回傳 dict（可能部分欄位為空）。
    """
    import openpyxl

    if isinstance(file_obj, (bytes, bytearray)):
        file_obj = io.BytesIO(file_obj)

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    merged_map = _build_merged_map(ws)

    def cv(r, c):
        return _cell_val(ws, r, c, merged_map)

    # 掃描前 10 列找關鍵字
    meta = {"model": "", "date": "", "customer": "", "qty": 0}

    for row in range(1, 11):
        for col in range(1, 20):
            v = cv(row, col)
            if not v:
                continue
            vn = v.upper()
            # 機種（通常在 'MD' 或 'ES' 開頭的儲存格旁邊）
            if re.match(r"^(MD|ES|PJ)\d", v) and not meta["model"]:
                meta["model"] = v
            # 數量
            if v.isdigit() and int(v) > 0 and not meta["qty"]:
                # 找 col label "本批數量" 附近的值
                label_left = cv(row, col - 1) + cv(row - 1, col)
                if "數量" in label_left or "批量" in label_left:
                    meta["qty"] = int(v)

    return meta
